import openpyxl


def getData():
    wk = openpyxl.load_workbook("E:/LoginDetails.xlsx")
    sh = wk['LoginSheet']
    r = sh.max_row
    c = sh.max_column
    dict = {}
    for i in range(1, r + 1):
        if sh.cell(row=i, column=1).value == "Testcase1":
            for j in range(2, c + 1):
                dict[sh.cell(row=1, column=j).value] = sh.cell(row=i, column=j).value

    """cell =sh.cell(row=1, column=2)
    r = sh.max_row
    li = []
    li1 = []
    for i in range(1, r+1):
        li1=[]
        un = sh.cell(i, 1)
        up = sh.cell(i, 2)
        li1.insert(0, un.value)
        li1.insert(1, up.value)
        li.insert(i-1, li1)
    print(li)
    return li"""
